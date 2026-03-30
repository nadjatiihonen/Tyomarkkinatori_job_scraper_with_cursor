from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "tyomarkkinatori_jobs.xlsx"


def looks_like_location(value: str) -> bool:
    s = (value or "").strip()
    if not s:
        return False
    low = s.lower()
    tokens = (
        "helsinki",
        "espoo",
        "vantaa",
        "tampere",
        "turku",
        "hämeenlinna",
        "koko suomi",
        "suomi",
        "etätyö",
        "etatyö",
    )
    if any(t in low for t in tokens):
        return True
    return " " not in s and len(s) <= 14


def title_has_company_hint(title: str) -> bool:
    """True if title likely contains company segment. / Tosi, jos otsikossa on todennäköisesti yritysosa."""
    t = (title or "").strip()
    if "," not in t:
        return False
    tail = t.split(",")[-1].strip().lower()
    if not tail:
        return False
    company_tokens = (" oy", " oyj", " ab", " ltd", " inc", " ky", " tmi", " ry")
    return any(tok in f" {tail}" for tok in company_tokens)


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: Excel file not found: {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=False)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wb.close()

    try:
        title_idx = headers.index("Tehtävänimike") + 1
        company_idx = headers.index("Yritys") + 1
    except ValueError:
        print("ERROR: Required columns Tehtävänimike / Yritys not found")
        return 1

    wb = load_workbook(XLSX, data_only=False)
    ws = wb.active

    total = 0
    empty_title = 0
    empty_company_required = 0
    location_like_company = 0

    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, title_idx).value
        company = ws.cell(r, company_idx).value
        if not title or not str(title).strip():
            continue
        total += 1
        if not str(title).strip():
            empty_title += 1
        c = "" if company is None else str(company).strip()
        if not c:
            if title_has_company_hint(str(title)):
                empty_company_required += 1
            continue
        if looks_like_location(c):
            location_like_company += 1

    wb.close()

    if total == 0:
        print("ERROR: No data rows found")
        return 1
    if empty_title > 0:
        print(f"ERROR: Found {empty_title} rows with empty Tehtävänimike")
        return 1
    if empty_company_required > 0:
        print(
            "ERROR: Yritys is empty in rows that likely contain company in title: "
            f"{empty_company_required}/{total}"
        )
        return 1
    if location_like_company > 0:
        print(
            f"ERROR: Yritys contains location-like values: {location_like_company}/{total}"
        )
        return 1

    print(
        "OK: data quality passed "
        f"(rows={total}, empty_required={empty_company_required}, location_like={location_like_company})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
