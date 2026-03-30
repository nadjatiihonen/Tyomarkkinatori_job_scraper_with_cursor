from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "tyomarkkinatori_jobs.xlsx"


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: Excel file not found: {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=False)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wb.close()

    expected_prefix = ["Tehtävänimike", "Yritys"]
    if headers[:2] != expected_prefix:
        print(f"ERROR: First two columns must be {expected_prefix}, got {headers[:2]}")
        return 1

    if len(headers) != 2:
        print(f"ERROR: Expected exactly 2 columns, got {len(headers)}: {headers}")
        return 1

    print("OK: schema is valid")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
