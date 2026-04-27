#!/usr/bin/env python3
"""
Collect jobs from tyomarkkinatori.fi into Excel: JSON API listing, sync, missing Yritys.
/ Työmarkkinatori.fi → Excel: JSON-API-listaus, synkronointi, puuttuvat Yritys-kentät.

Flow: (1) POST /api/jobpostingfulltext/search/v2/search (same filters as listing URL)
      (2) Excel sync + save
      (3) Playwright only if Yritys still empty (JSON-LD / label on detail page)
/ Rakenne: (1) virallinen hakurajapinta
           (2) Excel-synk + tallennus
           (3) Playwright vain puuttuvaa Yritys varten
"""
from __future__ import annotations

import os
import re
import sys
import tempfile
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any, Optional
from urllib.parse import parse_qs, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(line_buffering=True)

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, sync_playwright

# Paths and HTTP / Polku ja verkko
EXCEL_PATH = Path(__file__).resolve().parent / "tyomarkkinatori_jobs.xlsx"
BASE_DOMAIN = "https://tyomarkkinatori.fi"
# Listing template (used only to derive API filters). / Listaus-URL → API-suodattimet
LISTING_URL = (
    "https://tyomarkkinatori.fi/henkiloasiakkaat/avoimet-tyopaikat"
    "?in=25&or=CLOSING&p={p}&ps=30"
)
KIRJANPITO_LISTING_URL = (
    "https://tyomarkkinatori.fi/henkiloasiakkaat/avoimet-tyopaikat"
    "?q=kirjanpito&or=CLOSING&p={p}&ps=30"
)
KOULUTUS_LISTING_URL = (
    "https://tyomarkkinatori.fi/henkiloasiakkaat/koulutukset-ja-palvelut"
    "?q=rekry&m=0&y=0&pa=1&eo=3&s=All&rs=All&re"
)
API_SEARCH_URL = f"{BASE_DOMAIN}/api/jobpostingfulltext/search/v2/search"
TRAINING_GRAPHQL_URL = f"{BASE_DOMAIN}/api/employmentservicecatalogue/graphql"
JOB_PATH_PREFIX = "/henkiloasiakkaat/avoimet-tyopaikat"
TRAINING_PATH_PREFIX = "/henkiloasiakkaat/koulutukset-ja-palvelut/kurssi"
SHEET_CONFIGS: list[dict[str, str]] = [
    {"sheet_name": "IT", "listing_url": LISTING_URL},
    {"sheet_name": "Kirjanpito", "listing_url": KIRJANPITO_LISTING_URL},
]
KOULUTUS_SHEET_NAME = "koulutus"
KOULUTUS_COLUMNS = ["Ohjelma", "Järjestäjä", "Sijainti", "Kesto", "Haku paättyy", "Julkaistu"]

# Timeouts / Aikarajat
REQUEST_TIMEOUT_S = 60
PAGE_GOTO_TIMEOUT_MS = 45_000
CARD_PAGE_WAIT_MS = 1_500

# Save Excel after each successful Yritys update (crash safety).
SAVE_AFTER_EVERY_DETAIL_WRITE = True

# Max listing pages (safety). / Listasivujen yläraja
MAX_LISTING_PAGES = 600

# DataFrame columns / DataFrame-sarakkeet
DATA_COLUMNS = ["Linkki", "Tehtävänimike", "Yritys", "Työsuhde", "Työaika", "Julkaistu"]

# Company title suffix fallback / Otsikon Oy-tms.-fallback
COMPANY_SUFFIXES = (" oy", " oyj", " ab", " ltd", " ky", " tmi", " ry", " inc")
DEFAULT_CONTINUITY_LABELS = {
    "01": "toistaiseksi voimassa oleva",
    "02": "määräaikainen",
}
DEFAULT_WORKTIME_LABELS = {
    "01": "kokoaikainen",
    "02": "osa-aikainen",
}


def _ascii_fold(s: str) -> str:
    """Lowercase ASCII for Finnish city matching. / ASCII-pienet kirjaimet ää→a."""
    return (
        unicodedata.normalize("NFKD", (s or "").strip())
        .encode("ascii", "ignore")
        .decode()
        .lower()
        .strip()
    )


def _slugify(value: str) -> str:
    s = unicodedata.normalize("NFKD", (value or "").strip()).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def _training_token_from_url(url: str) -> str:
    s = str(url or "").strip()
    if not s:
        return ""
    m = re.search(
        r"/koulutukset-ja-palvelut/kurssi/([0-9a-fA-F-]{36})",
        s,
    )
    return (m.group(1) if m else "").lower()


# Major Finnish cities (exact match only, ASCII-folded). No "short word" heuristic.
# / Suuret kaupungit: vain täsmäosuma; ei "lyhyt sana = kaupunki" -heuristiikkaa.
_FI_CITY_NAMES = (
    "Helsinki",
    "Espoo",
    "Tampere",
    "Vantaa",
    "Turku",
    "Oulu",
    "Jyväskylä",
    "Lahti",
    "Kuopio",
    "Pori",
    "Kouvola",
    "Joensuu",
    "Lappeenranta",
    "Hämeenlinna",
    "Vaasa",
    "Seinäjoki",
    "Rovaniemi",
    "Mikkeli",
    "Kotka",
    "Salo",
    "Hyvinkää",
    "Porvoo",
    "Kajaani",
    "Rauma",
    "Lohja",
    "Järvenpää",
    "Kirkkonummi",
    "Kerava",
    "Tuusula",
    "Savonlinna",
    "Riihimäki",
    "Valkeakoski",
)

FI_MAJOR_CITIES_ASCII = frozenset(
    _ascii_fold(c) for c in _FI_CITY_NAMES if _ascii_fold(c)
)


# ---------------------------------------------------------------------------
# URLs / URLit
# ---------------------------------------------------------------------------


def canonical_job_url(url: str) -> str:
    """Stable key: no query string, domain always present. / Yhtenäinen avain."""
    s = str(url).strip()
    if not s or s.lower() == "nan":
        return ""
    s = s.split("?", 1)[0].rstrip("/")
    if s and not s.startswith("http"):
        s = f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    return s


def job_url_from_api_id(job_id: str) -> str:
    """Public job page URL from API id. / Ilmoituksen sivun URL."""
    jid = (job_id or "").strip()
    if not jid:
        return ""
    return canonical_job_url(f"{BASE_DOMAIN}{JOB_PATH_PREFIX}/{jid}")


# ---------------------------------------------------------------------------
# API: search body from LISTING_URL / Hakukysely listaus-URL:sta
# ---------------------------------------------------------------------------


def _search_request_body(
    page_number: int,
    listing_url_template: str,
) -> tuple[dict[str, Any], int]:
    """
    Build JSON body like the site widget (short query params → filters).
    Supported: q, in (iscoNotations), or (sorting), p, ps.
    / Rakenna sama pyyntö kuin sivusto: q, in, or, p, ps.
    """
    filled = listing_url_template.format(p=page_number)
    parsed = urlparse(filled)
    q = parse_qs(parsed.query, keep_blank_values=True)
    page_size = int((q.get("ps") or ["30"])[0])
    sorting = (q.get("or") or ["CLOSING"])[0]
    query_text = (q.get("q") or [""])[0]
    filters: dict[str, Any] = {}
    if q.get("in"):
        filters["iscoNotations"] = q["in"]
    body: dict[str, Any] = {
        "query": query_text,
        "filters": filters,
        "paging": {"pageNumber": page_number, "pageSize": page_size},
        "sorting": sorting,
    }
    return body, page_size


def _employer_name_from_api_item(item: dict[str, Any]) -> str:
    """employer.ownerName.{fi,sv,en} or employer.name. / Työnantajan nimi API-kentistä."""
    emp = item.get("employer") or {}
    on = emp.get("ownerName")
    if isinstance(on, dict):
        for lang in ("fi", "sv", "en"):
            v = (on.get(lang) or "").strip()
            if v:
                return v[:500]
        for v in on.values():
            v = (str(v) or "").strip()
            if v:
                return v[:500]
    n = (emp.get("name") or "").strip()
    if n:
        return n[:500]
    return ""


def _title_from_api_item(item: dict[str, Any]) -> str:
    t = item.get("title")
    if isinstance(t, dict):
        for lang in ("fi", "sv", "en"):
            v = (t.get(lang) or "").strip()
            if v:
                return v[:500]
        for v in t.values():
            v = (str(v) or "").strip()
            if v:
                return v[:500]
    return (str(t) if t else "").strip()[:500]


def _continuity_code_to_tyosuhde(code: str) -> str:
    """Map continuity code to contract type label. / Muunna jatkuvuuskoodi työsuhde-tekstiksi."""
    c = (code or "").strip()
    if not c:
        return ""
    if c.startswith("01"):
        return DEFAULT_CONTINUITY_LABELS["01"]
    if c.startswith("02"):
        return DEFAULT_CONTINUITY_LABELS["02"]
    return ""


def _worktime_code_to_tyoaika(code: str) -> str:
    """Map worktime code to worktime label. / Muunna työaikakoodi työaika-tekstiksi."""
    c = (code or "").strip()
    if not c:
        return ""
    if c.startswith("01"):
        return DEFAULT_WORKTIME_LABELS["01"]
    if c.startswith("02"):
        return DEFAULT_WORKTIME_LABELS["02"]
    return ""


def _iso_date_ymd(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    return s[:10]


def fetch_continuity_labels(
    session: requests.Session,
) -> dict[str, str]:
    """
    Load TYÖN_JATKUVUUS code labels from API.
    / Hae TYÖN_JATKUVUUS-koodien selitteet API:sta.
    """
    url = (
        f"{BASE_DOMAIN}/api/codes/v1/kopa/TY%C3%96N_JATKUVUUS/koodit"
        f"?voimassa={date.today().isoformat()}"
    )
    out: dict[str, str] = {}
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT_S)
        r.raise_for_status()
        payload = r.json()
    except Exception:
        return out

    if not isinstance(payload, list):
        return out
    for row in payload:
        if not isinstance(row, dict):
            continue
        code = str(row.get("tunnus") or "").strip()
        labels = row.get("selite") or []
        if not code or not isinstance(labels, list):
            continue
        fi = ""
        for item in labels:
            if not isinstance(item, dict):
                continue
            if (item.get("kielikoodi") or "").strip().lower() == "fi":
                fi = str(item.get("teksti") or "").strip()
                break
        if fi:
            # Normalize output to the two required classes.
            out[code] = _continuity_code_to_tyosuhde(code) or fi.lower()
    return out


def fetch_worktime_labels(
    session: requests.Session,
) -> dict[str, str]:
    """
    Load TYÖAIKA code labels from API.
    / Hae TYÖAIKA-koodien selitteet API:sta.
    """
    url = (
        f"{BASE_DOMAIN}/api/codes/v1/kopa/TY%C3%96AIKA/koodit"
        f"?voimassa={date.today().isoformat()}"
    )
    out: dict[str, str] = {}
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT_S)
        r.raise_for_status()
        payload = r.json()
    except Exception:
        return out

    if not isinstance(payload, list):
        return out
    for row in payload:
        if not isinstance(row, dict):
            continue
        code = str(row.get("tunnus") or "").strip()
        labels = row.get("selite") or []
        if not code or not isinstance(labels, list):
            continue
        fi = ""
        for item in labels:
            if not isinstance(item, dict):
                continue
            if (item.get("kielikoodi") or "").strip().lower() == "fi":
                fi = str(item.get("teksti") or "").strip()
                break
        if fi:
            out[code] = _worktime_code_to_tyoaika(code) or fi.lower()
    return out


def _tyosuhde_from_api_item(
    item: dict[str, Any],
    continuity_labels: dict[str, str],
) -> str:
    """Työsuhde from continuityOfWork code list. / Työsuhde continuityOfWork-koodista."""
    raw = item.get("continuityOfWork") or []
    if not isinstance(raw, list):
        raw = [raw]
    for code in raw:
        c = str(code or "").strip()
        if not c:
            continue
        mapped = continuity_labels.get(c) or _continuity_code_to_tyosuhde(c)
        if mapped:
            return mapped[:120]
    return ""


def _tyoaika_from_api_item(
    item: dict[str, Any],
    worktime_labels: dict[str, str],
) -> str:
    """Työaika from workTime code. / Työaika workTime-koodista."""
    code = str(item.get("workTime") or "").strip()
    if not code:
        return ""
    mapped = worktime_labels.get(code) or _worktime_code_to_tyoaika(code)
    return mapped[:120] if mapped else ""


def job_row_from_api_item(
    item: dict[str, Any],
    continuity_labels: dict[str, str],
    worktime_labels: dict[str, str],
) -> dict[str, str]:
    """One row: Linkki, Tehtävänimike, Yritys from API hit. / Yksi rivi API:sta."""
    jid = (item.get("id") or "").strip()
    link = job_url_from_api_id(jid)
    title = _title_from_api_item(item) or "-"
    yritys = clean_company_name(_employer_name_from_api_item(item))
    tyosuhde = _tyosuhde_from_api_item(item, continuity_labels)
    tyoaika = _tyoaika_from_api_item(item, worktime_labels)
    julkaistu = _iso_date_ymd(item.get("publishDate") or item.get("created"))
    if looks_like_location(yritys):
        yritys = ""
    return {
        "Linkki": link,
        "Tehtävänimike": title,
        "Yritys": yritys,
        "Työsuhde": tyosuhde,
        "Työaika": tyoaika,
        "Julkaistu": julkaistu,
    }


def fetch_all_listings_api(
    listing_url_template: str,
    session: Optional[requests.Session] = None,
) -> list[dict]:
    """
    All jobs via search API (fast, structured employer names).
    / Kaikki ilmoitukset hakurajapinnasta.
    """
    sess = session or requests.Session()
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Referer": f"{BASE_DOMAIN}/",
    }
    all_jobs: list[dict] = []
    page_num = 0
    continuity_labels = fetch_continuity_labels(sess)
    worktime_labels = fetch_worktime_labels(sess)

    while page_num < MAX_LISTING_PAGES:
        body, _ = _search_request_body(page_num, listing_url_template)
        print(f"Haetaan API-sivu {page_num + 1}...", flush=True)
        try:
            r = sess.post(
                API_SEARCH_URL,
                json=body,
                headers=headers,
                timeout=REQUEST_TIMEOUT_S,
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"API-virhe (sivu {page_num + 1}): {e}", flush=True)
            break

        content = data.get("content") or []
        if not content:
            break

        for item in content:
            if isinstance(item, dict):
                all_jobs.append(
                    job_row_from_api_item(item, continuity_labels, worktime_labels)
                )

        print(
            f"Löytyi {len(content)} ilmoitusta (yhteensä {len(all_jobs)}).",
            flush=True,
        )
        if len(content) < body["paging"]["pageSize"]:
            break
        page_num += 1

    return all_jobs


def _value_for_language(items: list[dict[str, Any]], lang: str = "fi") -> str:
    for item in items or []:
        if not isinstance(item, dict):
            continue
        if (item.get("language") or "").strip().lower() == lang:
            v = str(item.get("value") or "").strip()
            if v:
                return v
    for item in items or []:
        if not isinstance(item, dict):
            continue
        v = str(item.get("value") or "").strip()
        if v:
            return v
    return ""


def _build_training_where(search_term: str) -> dict[str, Any]:
    return {
        "and": [
            {
                "or": [
                    {"names": {"some": {"value": {"contains": search_term}}}},
                    {"additionalInformation": {"some": {"value": {"contains": search_term}}}},
                    {"code": {"eq": search_term}},
                    {
                        "serviceOffering": {
                            "or": [
                                {
                                    "organizations": {
                                        "some": {
                                            "organization": {
                                                "names": {
                                                    "some": {"value": {"contains": search_term}}
                                                }
                                            }
                                        }
                                    }
                                },
                                {
                                    "service": {
                                        "or": [
                                            {"names": {"some": {"value": {"contains": search_term}}}},
                                            {
                                                "description": {
                                                    "some": {"value": {"contains": search_term}}
                                                }
                                            },
                                            {"summary": {"some": {"value": {"contains": search_term}}}},
                                            {
                                                "serviceClasses": {
                                                    "some": {
                                                        "serviceClass": {
                                                            "name": {
                                                                "some": {
                                                                    "value": {"contains": search_term}
                                                                }
                                                            }
                                                        }
                                                    }
                                                }
                                            },
                                            {"keywords": {"some": {"value": {"contains": search_term}}}},
                                        ]
                                    }
                                },
                            ]
                        }
                    },
                ]
            },
            {
                "or": [
                    {
                        "serviceOffering": {
                            "service": {
                                "category": {
                                    "code": {
                                        "in": [
                                            "PK2",
                                            "PK2.1",
                                            "PK2.2",
                                            "PK2.3",
                                            "PK2.4",
                                            "PK7",
                                            "PK7.1",
                                            "PK7.2",
                                            "PK7.3",
                                            "PK7.4",
                                            "PK7.4.1",
                                            "PK7.4.2",
                                            "PK7.4.3",
                                            "PK13",
                                        ]
                                    }
                                }
                            }
                        }
                    }
                ]
            },
            {"serviceOffering": {"service": {"origin": {"nin": ["Ecosystem"]}}}},
        ],
        "status": {"eq": "Published"},
        "serviceOffering": {
            "service": {"moderationStatus": {"eq": "Active"}, "status": {"eq": "Published"}}
        },
    }


def _training_row_from_item(item: dict[str, Any]) -> dict[str, str]:
    token = str(item.get("token") or "").strip()
    ohjelma = _value_for_language(item.get("names") or []) or "-"
    slug = _slugify(ohjelma) or token
    link = (
        canonical_job_url(f"{BASE_DOMAIN}{TRAINING_PATH_PREFIX}/{token}/{slug}")
        if token
        else ""
    )

    jarjestaja = ""
    for org_entry in ((item.get("serviceOffering") or {}).get("organizations") or []):
        org = (org_entry or {}).get("organization") or {}
        name = _value_for_language(org.get("names") or [])
        if name:
            jarjestaja = name
            break

    city_names: list[str] = []
    for municipality in ((item.get("area") or {}).get("municipalities") or []):
        name = _value_for_language((municipality or {}).get("names") or [])
        if name and name not in city_names:
            city_names.append(name)
    sijainti = ", ".join(city_names)

    start = str(item.get("startDate") or "").strip()[:10]
    end = str(item.get("endDate") or "").strip()[:10]
    if start and end:
        kesto = f"{start} - {end}"
    elif start or end:
        kesto = start or end
    else:
        kesto = _value_for_language(item.get("implementationDurationDescription") or [])
    haku_paattyy = str(item.get("publicationEndDate") or "").strip()[:10]
    julkaistu = _iso_date_ymd(item.get("publicationStartDate") or item.get("created"))

    return {
        "Linkki": link,
        "Ohjelma": ohjelma[:500],
        "Järjestäjä": jarjestaja[:300],
        "Sijainti": sijainti[:300],
        "Kesto": kesto[:120],
        "Haku paättyy": haku_paattyy,
        "Julkaistu": julkaistu,
    }


def fetch_koulutus_listings_api(
    listing_url: str,
    session: Optional[requests.Session] = None,
) -> list[dict[str, str]]:
    parsed = urlparse(listing_url)
    params = parse_qs(parsed.query, keep_blank_values=True)
    search_term = (params.get("q") or [""])[0].strip()
    page_size = 12
    offset = 0
    out: list[dict[str, str]] = []
    sess = session or requests.Session()

    query = """
query GetPublicServiceImplementationsPaginated($where: ServiceImplementationFilterInput!, $pageSize: Int!, $after: Int, $orderBy: [OrderByInput]) {
  serviceImplementation {
    allServiceImplementationsPaginated(skip: $after, take: $pageSize, where: $where, orderBy: $orderBy) {
      pageInfo { hasNextPage }
      items {
        token
        publicationStartDate
        names { language value }
        publicationEndDate
        startDate
        endDate
        implementationDurationDescription { language value }
        area { municipalities { names { language value } } }
        serviceOffering { organizations { organization { names { language value type } } } }
      }
    }
  }
}
""".strip()

    while True:
        variables = {
            "where": _build_training_where(search_term),
            "pageSize": page_size,
            "after": offset,
            "orderBy": [
                {"field": "serviceImplementationName", "language": "fi", "order": "Ascending"}
            ],
        }
        payload = {
            "query": query,
            "variables": variables,
            "operationName": "GetPublicServiceImplementationsPaginated",
        }
        r = sess.post(TRAINING_GRAPHQL_URL, json=payload, timeout=REQUEST_TIMEOUT_S)
        r.raise_for_status()
        data = r.json()
        page = (
            (data.get("data") or {})
            .get("serviceImplementation", {})
            .get("allServiceImplementationsPaginated", {})
        )
        items = page.get("items") or []
        if not items:
            break
        for item in items:
            if isinstance(item, dict):
                out.append(_training_row_from_item(item))
        if not (page.get("pageInfo") or {}).get("hasNextPage"):
            break
        offset += page_size

    return out


# ---------------------------------------------------------------------------
# Job page: Yritys fallback (Playwright) / Ilmoitussivu: varayritys
# ---------------------------------------------------------------------------


def clean_company_name(value: str) -> str:
    """Return plausible company string or empty. / Kelvollinen yritysnimi tai tyhjä."""
    s = (value or "").strip()
    if not s or len(s) > 80:
        return ""
    low = s.lower()
    if any(
        t in low
        for t in (
            " toimii ",
            " yrityksessä",
            " tehtävässä",
            " rooli ",
            " tiimi ",
        )
    ):
        return ""
    return s


def looks_like_location(value: str) -> bool:
    """
    True only for exact major-city match or obvious multi-city location strings.
    No generic "short single word" rule.
    / Vain täsmäkaupunki tai selvä monen kaupungin -merkkijono.
    """
    s = (value or "").strip()
    if not s:
        return False
    key = _ascii_fold(s)
    if key in FI_MAJOR_CITIES_ASCII:
        return True
    low = s.lower()
    remote_exact = (
        "koko suomi",
        "koko suomessa",
        "100% etätyö",
        "100% etatyö",
        "etätyö",
        "etatyö",
    )
    if low in remote_exact:
        return True
    if re.search(r"\s+or\s+|\s+tai\s+", low):
        parts = [p.strip() for p in re.split(r"\s+or\s+|\s+tai\s+", low) if p.strip()]
        if len(parts) >= 2 and all(
            _ascii_fold(p) in FI_MAJOR_CITIES_ASCII for p in parts
        ):
            return True
    return False


def company_from_title_fallback(title: str) -> str:
    """Last comma segment if legal suffix (Oy, …). / Viimeinen pilkkuerotettu vain jos Oy-tms."""
    t = (title or "").strip()
    if not t or "," not in t:
        return ""
    cand = t.split(",")[-1].strip()
    cand = clean_company_name(cand)
    if not cand or looks_like_location(cand):
        return ""
    low = f" {cand.lower()}"
    if any(tok in low for tok in COMPANY_SUFFIXES):
        return cand
    return ""


def extract_company_json_ld_and_label(page: Page) -> str:
    """JSON-LD hiringOrganization, then Yritys label before location. / JSON-LD + Yritys-teksti."""
    try:
        return (
            page.evaluate(
                """
            () => {
                const clean = (v) => (v || '').toString().trim();
                const ldScripts = Array.from(
                    document.querySelectorAll('script[type="application/ld+json"]')
                );
                for (const s of ldScripts) {
                    try {
                        const raw = clean(s.textContent);
                        if (!raw) continue;
                        const data = JSON.parse(raw);
                        const items = Array.isArray(data) ? data : [data];
                        for (const item of items) {
                            const org = item && item.hiringOrganization;
                            const name = clean(org && org.name);
                            if (name) return name;
                        }
                    } catch (_) {}
                }
                const bodyText = clean(document.body && document.body.innerText);
                if (bodyText) {
                    const headEnd = bodyText.toLowerCase()
                        .indexOf('työpaikan sijainti');
                    const block = headEnd >= 0
                        ? bodyText.slice(0, headEnd)
                        : bodyText;
                    const m = block.match(/Yritys\\s*[:\\s\\u00a0]+([^\\n]+)/i);
                    if (m && m[1]) return clean(m[1]);
                }
                return '';
            }
            """
            )
            or ""
        ).strip()
    except Exception:
        return ""


def fetch_yritys_from_job_page(page: Page, url: str, title_hint: str = "") -> str:
    """Open job page; Yritys from JSON-LD / label, else title suffix. / Avaa sivu, poimi Yritys."""
    s = (url or "").strip()
    full = s if s.startswith("http") else f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    try:
        page.goto(full, wait_until="load", timeout=PAGE_GOTO_TIMEOUT_MS)
        page.wait_for_timeout(CARD_PAGE_WAIT_MS)
    except Exception:
        return ""

    y = clean_company_name(extract_company_json_ld_and_label(page))
    if looks_like_location(y):
        y = ""
    if not y:
        y = company_from_title_fallback(title_hint)
    return y


# ---------------------------------------------------------------------------
# Excel / Excel
# ---------------------------------------------------------------------------


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure Linkki, Tehtävänimike, Yritys. / Varmista sarakkeet."""
    for col in DATA_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def _get_or_create_worksheet(wb, sheet_name: str):
    wanted = (sheet_name or "").strip().lower()
    for existing in wb.sheetnames:
        if existing.strip().lower() == wanted:
            return wb[existing]
    return wb.create_sheet(title=sheet_name)


def extract_urls_and_titles_from_excel(
    path: Path,
    n_rows: int,
    sheet_name: str,
) -> tuple[list[str], list[str]]:
    """URLs and titles from Tehtävänimike hyperlinks. / URL ja otsikko linkeistä."""
    urls: list[str] = []
    titles: list[str] = []
    try:
        wb = load_workbook(path, data_only=False)
        ws = _get_or_create_worksheet(wb, sheet_name)
        col_idx = 1
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Tehtävänimike":
                col_idx = c
                break
        for r in range(2, ws.max_row + 1):
            if len(urls) >= n_rows:
                break
            cell = ws.cell(row=r, column=col_idx)
            href = None
            title = ""
            val = cell.value
            h = getattr(cell, "hyperlink", None)
            if h:
                href = getattr(h, "target", None) or getattr(h, "location", None)
            if isinstance(val, str) and "HYPERLINK" in val:
                m_url = re.search(r'HYPERLINK\s*\(\s*"((?:[^"]|"")+)"', val, re.I)
                if m_url:
                    href = href or m_url.group(1).replace('""', '"')
                m_title = re.search(
                    r'HYPERLINK\s*\(\s*"(?:[^"]|"")*"\s*,\s*"((?:[^"]|"")*)"\s*\)',
                    val,
                    re.I,
                )
                if m_title:
                    title = m_title.group(1).replace('""', '"').strip()
            elif isinstance(val, str) and val.strip() and not val.strip().startswith("="):
                title = val.strip()[:500]
            urls.append(str(href or "").strip())
            titles.append(title)
        wb.close()
    except Exception:
        pass
    return urls, titles


def sync_dataframe(
    jobs_from_web: list[dict],
    sheet_name: str,
) -> tuple[pd.DataFrame, int, int, set[str]]:
    """DataFrame = only jobs on site. / Taulukko = vain sivulla olevat."""
    web_links = {canonical_job_url(j["Linkki"]) for j in jobs_from_web}
    web_links.discard("")

    if not EXCEL_PATH.exists():
        df = ensure_columns(pd.DataFrame(jobs_from_web))
        return df, len(jobs_from_web), 0, set()

    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    except Exception as e:
        print(f"Excelin lukuvirhe: {e}", flush=True)
        df = ensure_columns(pd.DataFrame(jobs_from_web))
        return df, len(jobs_from_web), 0, set()

    df = ensure_columns(df)
    if "Tehtävänimike" in df.columns and df["Tehtävänimike"].dtype != object:
        df["Tehtävänimike"] = df["Tehtävänimike"].astype(object)
    for col in ("Yritys", "Työsuhde", "Työaika"):
        if col in df.columns and df[col].dtype != object:
            df[col] = df[col].astype(object)

    urls, titles = extract_urls_and_titles_from_excel(EXCEL_PATH, len(df), sheet_name)
    if len(urls) < len(df):
        urls.extend([""] * (len(df) - len(urls)))
        titles.extend([""] * (len(df) - len(titles)))
    df["Linkki"] = urls[: len(df)]

    def is_empty_title(x: object) -> bool:
        if pd.isna(x):
            return True
        s = str(x).strip().lower()
        return not s or s == "nan"

    for i, tit in enumerate(titles[: len(df)]):
        if tit and tit.lower() != "nan" and is_empty_title(df.at[i, "Tehtävänimike"]):
            df.at[i, "Tehtävänimike"] = tit

    excel_links = {
        canonical_job_url(x) for x in df["Linkki"].astype(str).unique() if canonical_job_url(x)
    }
    new_links = web_links - excel_links
    removed_links = excel_links - web_links

    df = df[~df["Linkki"].astype(str).apply(canonical_job_url).isin(removed_links)].copy()
    new_rows = [j for j in jobs_from_web if canonical_job_url(j["Linkki"]) in new_links]
    if new_rows:
        new_df = pd.DataFrame(new_rows)
        for c in df.columns:
            if c not in new_df.columns:
                new_df[c] = ""
        new_df = new_df.reindex(columns=df.columns, fill_value="")
        df = pd.concat([df, new_df], ignore_index=True)

    df = df[df["Linkki"].astype(str).apply(canonical_job_url).isin(web_links)].copy()
    df = df.drop_duplicates(subset=["Linkki"], keep="first")

    link_to_title = {
        canonical_job_url(j["Linkki"]): j.get("Tehtävänimike", "") for j in jobs_from_web
    }
    for i in df.index:
        t = link_to_title.get(canonical_job_url(str(df.at[i, "Linkki"])), "")
        if t and str(t).strip().lower() != "nan":
            # Keep title aligned with the current posting behind this URL.
            df.at[i, "Tehtävänimike"] = str(t)
        elif is_empty_title(df.at[i, "Tehtävänimike"]):
            df.at[i, "Tehtävänimike"] = "-"

    link_to_yritys: dict[str, str] = {}
    link_to_tyosuhde: dict[str, str] = {}
    link_to_tyoaika: dict[str, str] = {}
    link_to_julkaistu: dict[str, str] = {}
    for j in jobs_from_web:
        lk = canonical_job_url(j["Linkki"])
        if not lk:
            continue
        raw_y = str(j.get("Yritys", "") or "").strip()
        if not raw_y:
            continue
        y = clean_company_name(raw_y)
        if y and not looks_like_location(y):
            link_to_yritys[lk] = y
        ts = str(j.get("Työsuhde", "") or "").strip()
        if ts:
            link_to_tyosuhde[lk] = ts
        ta = str(j.get("Työaika", "") or "").strip()
        if ta:
            link_to_tyoaika[lk] = ta
        jul = str(j.get("Julkaistu", "") or "").strip()
        if jul:
            link_to_julkaistu[lk] = jul
    for i in df.index:
        y = link_to_yritys.get(canonical_job_url(str(df.at[i, "Linkki"])), "")
        if y:
            df.at[i, "Yritys"] = y
        ts = link_to_tyosuhde.get(canonical_job_url(str(df.at[i, "Linkki"])), "")
        if ts:
            df.at[i, "Työsuhde"] = ts
        ta = link_to_tyoaika.get(canonical_job_url(str(df.at[i, "Linkki"])), "")
        if ta:
            df.at[i, "Työaika"] = ta
        jul = link_to_julkaistu.get(canonical_job_url(str(df.at[i, "Linkki"])), "")
        if jul:
            df.at[i, "Julkaistu"] = jul

    order = {canonical_job_url(j["Linkki"]): i for i, j in enumerate(jobs_from_web)}
    df["_ord"] = df["Linkki"].astype(str).apply(
        lambda x: order.get(canonical_job_url(x), 999)
    )
    df = df.sort_values("_ord").drop(columns=["_ord"])
    df.reset_index(drop=True, inplace=True)
    return df, len(new_rows), len(removed_links), removed_links


def row_has_valid_yritys(df: pd.DataFrame, i: int) -> bool:
    """Non-empty Yritys and not classified as location. / Yritys ok."""
    if "Yritys" not in df.columns:
        return False
    v = df.at[i, "Yritys"]
    if v is None or not str(v).strip():
        return False
    return not looks_like_location(str(v))


def get_url_from_hyperlink_cell(cell) -> Optional[str]:
    h = getattr(cell, "hyperlink", None)
    if h:
        return getattr(h, "target", None) or getattr(h, "location", None)
    if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK("):
        m = re.search(r'=HYPERLINK\s*\(\s*"((?:[^"]|"")+)"', cell.value)
        if m:
            return m.group(1).replace('""', '"')
    return None


def find_title_column_index(ws) -> int:
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Tehtävänimike":
            return c
    return 1


def save_workbook_atomic(wb, path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
    os.close(fd)
    tmp_path = Path(tmp)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
        raise
    finally:
        wb.close()


def save_excel(df: pd.DataFrame, sheet_name: str) -> None:
    df = ensure_columns(df)
    display_cols = ["Tehtävänimike", "Yritys", "Työsuhde", "Työaika", "Julkaistu"]

    if not EXCEL_PATH.exists():
        df_out = df.reindex(columns=display_cols, fill_value="")
        df_out = df_out[display_cols]
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name=sheet_name)
        apply_hyperlinks_new_file(EXCEL_PATH, df["Linkki"].astype(str).tolist(), sheet_name)
        return

    wb = load_workbook(EXCEL_PATH)
    ws = _get_or_create_worksheet(wb, sheet_name)
    title_col = find_title_column_index(ws)

    link_to_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=title_col)
        url = get_url_from_hyperlink_cell(cell)
        if url:
            nu = canonical_job_url(url)
            if nu:
                link_to_row[nu] = r

    wanted = {canonical_job_url(str(row.get("Linkki", ""))) for _, row in df.iterrows()}
    wanted.discard("")
    keep_rows = {r for url, r in link_to_row.items() if url in wanted}
    to_delete = sorted(
        (r for r in range(2, ws.max_row + 1) if r not in keep_rows),
        reverse=True,
    )
    if to_delete:
        print(f"Poistetaan {len(to_delete)} riviä (vanhentuneet + tyhjät).", flush=True)
    for row_idx in to_delete:
        ws.delete_rows(row_idx, 1)
        for k in list(link_to_row):
            if link_to_row[k] > row_idx:
                link_to_row[k] -= 1
            elif link_to_row[k] == row_idx:
                del link_to_row[k]

    col_indices: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h in display_cols:
            col_indices[str(h)] = c
    for col in display_cols:
        if col not in col_indices:
            new_c = ws.max_column + 1
            ws.cell(row=1, column=new_c, value=col)
            col_indices[col] = new_c

    def write_display_row(target_row: int, row_data: pd.Series, link_url: str) -> None:
        for col in display_cols:
            if col in col_indices and col in row_data:
                val = (
                    row_data[col]
                    if col != "Tehtävänimike"
                    else row_data.get("Tehtävänimike", "-")
                )
                if pd.notna(val) and str(val).strip():
                    ws.cell(row=target_row, column=col_indices[col]).value = str(val)[:500]
                else:
                    ws.cell(row=target_row, column=col_indices[col]).value = None
        set_hyperlink_cell(
            ws,
            target_row,
            title_col,
            link_url,
            display=str(row_data.get("Tehtävänimike", "-")),
        )

    for _, row in df.iterrows():
        link_raw = str(row.get("Linkki", "")).strip()
        if not link_raw or link_raw == "nan":
            continue
        link = canonical_job_url(link_raw)
        if not link:
            continue

        if link in link_to_row:
            write_display_row(link_to_row[link], row, link)
            continue

        new_row = ws.max_row + 1
        write_display_row(new_row, row, link)
        link_to_row[link] = new_row

    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=title_col).value:
            last_data_row = r
    while ws.max_row > last_data_row:
        ws.delete_rows(ws.max_row, 1)

    save_workbook_atomic(wb, EXCEL_PATH)
    print(f"Tallennettu: {last_data_row - 1} riviä.", flush=True)


def apply_hyperlinks_new_file(path: Path, urls: list[str], sheet_name: str) -> None:
    try:
        wb = load_workbook(path)
        ws = _get_or_create_worksheet(wb, sheet_name)
        title_col = find_title_column_index(ws)
        for r in range(2, min(ws.max_row + 1, len(urls) + 2)):
            idx = r - 2
            if idx < len(urls) and urls[idx]:
                set_hyperlink_cell(ws, r, title_col, urls[idx])
        save_workbook_atomic(wb, path)
    except Exception:
        pass


def set_hyperlink_cell(
    ws,
    row: int,
    col: int,
    url: str,
    display: Optional[str] = None,
) -> None:
    s = str(url).strip()
    if ".fihenkilo" in s:
        s = s.replace(".fihenkilo", ".fi/henkilo", 1)
    elif not s.startswith("http"):
        s = f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    cell = ws.cell(row=row, column=col)
    raw = display if display is not None else cell.value
    label = (
        "-"
        if (raw is None or str(raw).strip().lower() in ("", "nan"))
        else str(raw)
    )
    cell.value = label
    cell.hyperlink = s


def apply_hyperlinks_to_worksheet(ws, df: pd.DataFrame, title_col: int) -> None:
    linkki_to_row_data: dict[str, Any] = {}
    for _, r in df.iterrows():
        lk = canonical_job_url(str(r.get("Linkki", "")))
        if lk:
            linkki_to_row_data[lk] = r
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=title_col)
        url = get_url_from_hyperlink_cell(cell)
        if url:
            nurl = canonical_job_url(url)
            if nurl in linkki_to_row_data:
                row_data = linkki_to_row_data[nurl]
                disp = row_data.get("Tehtävänimike")
                disp = "-" if (pd.isna(disp) or not str(disp).strip()) else str(disp)
                set_hyperlink_cell(ws, r, title_col, url, display=disp)


def save_koulutus_excel(df: pd.DataFrame, sheet_name: str) -> None:
    for col in ["Linkki"] + KOULUTUS_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    wb = load_workbook(EXCEL_PATH) if EXCEL_PATH.exists() else Workbook()
    ws = _get_or_create_worksheet(wb, sheet_name)

    # Remove duplicate tabs with same name but different case/spacing.
    canonical = (sheet_name or "").strip().lower()
    ws_title = ws.title
    dup_titles = [
        t
        for t in wb.sheetnames
        if t != ws_title and t.strip().lower() == canonical
    ]
    for title in dup_titles:
        wb.remove(wb[title])

    # Remove default empty sheet when it was auto-created.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        default_ws = wb["Sheet"]
        if default_ws.max_row <= 1 and default_ws.max_column <= 1 and not default_ws["A1"].value:
            wb.remove(default_ws)

    if ws.max_row < 1 or not ws.cell(row=1, column=1).value:
        for i, header in enumerate(KOULUTUS_COLUMNS, start=1):
            ws.cell(row=1, column=i, value=header)
    else:
        existing_headers: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if isinstance(val, str) and val.strip():
                existing_headers[val.strip()] = c
        for header in KOULUTUS_COLUMNS:
            if header not in existing_headers:
                new_c = ws.max_column + 1
                ws.cell(row=1, column=new_c, value=header)
                existing_headers[header] = new_c

    koulutus_col_idx: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val in KOULUTUS_COLUMNS:
            koulutus_col_idx[str(val)] = c

    wanted_rows: dict[str, dict[str, str]] = {}
    wanted_order: list[str] = []
    for _, row in df.iterrows():
        token = _training_token_from_url(str(row.get("Linkki", "")))
        if not token or token in wanted_rows:
            continue
        wanted_rows[token] = {
            "Ohjelma": str(row.get("Ohjelma", "") or "-")[:500],
            "Järjestäjä": str(row.get("Järjestäjä", "") or "")[:300],
            "Sijainti": str(row.get("Sijainti", "") or "")[:300],
            "Kesto": str(row.get("Kesto", "") or "")[:120],
            "Haku paättyy": str(row.get("Haku paättyy", "") or "")[:30],
            "Julkaistu": str(row.get("Julkaistu", "") or "")[:30],
            "Linkki": canonical_job_url(str(row.get("Linkki", "") or "")),
        }
        wanted_order.append(token)

    existing: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        url = get_url_from_hyperlink_cell(cell) or ""
        token = _training_token_from_url(url)
        if token:
            existing[token] = r

    to_delete = sorted(
        (row_idx for tok, row_idx in existing.items() if tok not in wanted_rows),
        reverse=True,
    )
    for row_idx in to_delete:
        ws.delete_rows(row_idx, 1)
        for tok in list(existing):
            if existing[tok] > row_idx:
                existing[tok] -= 1
            elif existing[tok] == row_idx:
                del existing[tok]

    for token in wanted_order:
        row = wanted_rows[token]
        new_row = existing[token] if token in existing else (ws.max_row + 1)
        ws.cell(row=new_row, column=koulutus_col_idx["Ohjelma"], value=row["Ohjelma"])
        ws.cell(row=new_row, column=koulutus_col_idx["Järjestäjä"], value=row["Järjestäjä"])
        ws.cell(row=new_row, column=koulutus_col_idx["Sijainti"], value=row["Sijainti"])
        ws.cell(row=new_row, column=koulutus_col_idx["Kesto"], value=row["Kesto"])
        ws.cell(row=new_row, column=koulutus_col_idx["Haku paättyy"], value=row["Haku paättyy"])
        ws.cell(row=new_row, column=koulutus_col_idx["Julkaistu"], value=row["Julkaistu"])
        if row["Linkki"]:
            ws.cell(row=new_row, column=koulutus_col_idx["Ohjelma"]).hyperlink = row["Linkki"]
        existing[token] = new_row

    save_workbook_atomic(wb, EXCEL_PATH)
    print(f"{sheet_name}: tallennettu {max(0, ws.max_row - 1)} riviä.", flush=True)


def sync_koulutus_dataframe(
    rows_from_web: list[dict[str, str]],
    sheet_name: str,
) -> tuple[pd.DataFrame, int, int]:
    web_rows: list[dict[str, str]] = []
    web_tokens: list[str] = []
    web_token_set: set[str] = set()
    for row in rows_from_web:
        token = _training_token_from_url(str(row.get("Linkki", "")))
        if not token or token in web_token_set:
            continue
        web_token_set.add(token)
        web_tokens.append(token)
        web_rows.append(row)

    if not EXCEL_PATH.exists():
        return pd.DataFrame(web_rows), len(web_rows), 0

    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame(web_rows), len(web_rows), 0

    for col in ["Linkki"] + KOULUTUS_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[["Linkki"] + KOULUTUS_COLUMNS].copy()

    urls, titles = extract_urls_and_titles_from_excel(EXCEL_PATH, len(df), sheet_name)
    if len(urls) < len(df):
        urls.extend([""] * (len(df) - len(urls)))
    if len(titles) < len(df):
        titles.extend([""] * (len(df) - len(titles)))
    df["Linkki"] = urls[: len(df)]
    for i, title in enumerate(titles[: len(df)]):
        if title and (not str(df.at[i, "Ohjelma"]).strip() or str(df.at[i, "Ohjelma"]).strip() == "nan"):
            df.at[i, "Ohjelma"] = title

    df["__token"] = df["Linkki"].astype(str).apply(_training_token_from_url)
    df = df[df["__token"] != ""].copy()
    excel_token_set = set(df["__token"].astype(str).tolist())
    new_tokens = web_token_set - excel_token_set
    removed_tokens = excel_token_set - web_token_set

    if removed_tokens:
        df = df[~df["__token"].isin(removed_tokens)].copy()

    new_rows = [
        r for r in web_rows if _training_token_from_url(str(r.get("Linkki", ""))) in new_tokens
    ]
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    token_to_published = {
        _training_token_from_url(str(r.get("Linkki", ""))): str(r.get("Julkaistu", "") or "")
        for r in web_rows
    }
    for i in df.index:
        tok = str(df.at[i, "__token"])
        if tok in token_to_published:
            df.at[i, "Julkaistu"] = token_to_published[tok]

    # Keep existing rows mostly unchanged; only publication date is refreshed.
    df["__token"] = df["Linkki"].astype(str).apply(_training_token_from_url)

    order = {token: idx for idx, token in enumerate(web_tokens)}
    df["_ord"] = df["__token"].astype(str).apply(lambda x: order.get(x, 999999))
    df = (
        df.sort_values("_ord")
        .drop(columns=["_ord"])
        .drop_duplicates(subset=["__token"], keep="first")
    )
    df = df.drop(columns=["__token"])
    df.reset_index(drop=True, inplace=True)
    return df, len(new_rows), len(removed_tokens)


def fill_missing_yritys_with_browser(
    page: Page,
    df: pd.DataFrame,
    sheet_name: str,
) -> None:
    """Playwright only for rows still missing Yritys. / Selain vain puuttuville."""
    if "Yritys" in df.columns:
        for i in range(len(df)):
            if looks_like_location(str(df.at[i, "Yritys"])):
                df.at[i, "Yritys"] = ""

    n = len(df)
    skipped = 0
    for i in range(n):
        link = df.at[i, "Linkki"]
        if pd.isna(link) or not str(link).strip():
            continue
        if row_has_valid_yritys(df, i):
            skipped += 1
            continue

        print(f"Haetaan Yritys sivulta {i + 1}/{n}...", flush=True)
        try:
            title_hint = (
                str(df.at[i, "Tehtävänimike"])
                if "Tehtävänimike" in df.columns
                else ""
            )
            val = fetch_yritys_from_job_page(page, str(link).strip(), title_hint=title_hint)
            if val:
                df.at[i, "Yritys"] = val
                if SAVE_AFTER_EVERY_DETAIL_WRITE:
                    save_excel(df, sheet_name)
        except Exception as e:
            print(f"Rivi {i + 1}/{n} ohitettu: {e}", flush=True)

    if skipped:
        print(f"Ohitettiin {skipped} riviä (Yritys jo API:sta / kelvollinen).", flush=True)


def needs_browser_for_yritys(df: pd.DataFrame) -> bool:
    for i in range(len(df)):
        if not row_has_valid_yritys(df, i):
            lk = df.at[i, "Linkki"]
            if pd.notna(lk) and str(lk).strip():
                return True
    return False


# ---------------------------------------------------------------------------
# Entry point / Käynnistys
# ---------------------------------------------------------------------------


def main() -> None:
    print("Työmarkkinatori -synkronointi alkaa.", flush=True)
    dfs: dict[str, pd.DataFrame] = {}

    def _fetch_jobs_for_sheet(cfg: dict[str, str]) -> tuple[str, list[dict]]:
        sheet_name = cfg["sheet_name"]
        with requests.Session() as http:
            jobs = fetch_all_listings_api(cfg["listing_url"], http)
        return sheet_name, jobs

    def _fetch_koulutus_for_sheet() -> tuple[str, list[dict[str, str]]]:
        with requests.Session() as http:
            rows = fetch_koulutus_listings_api(KOULUTUS_LISTING_URL, http)
        return KOULUTUS_SHEET_NAME, rows

    try:
        jobs_by_sheet: dict[str, list[dict]] = {}
        with ThreadPoolExecutor(max_workers=len(SHEET_CONFIGS) + 1) as executor:
            futures = [executor.submit(_fetch_jobs_for_sheet, cfg) for cfg in SHEET_CONFIGS]
            futures.append(executor.submit(_fetch_koulutus_for_sheet))
            for future in as_completed(futures):
                sheet_name, jobs = future.result()
                jobs_by_sheet[sheet_name] = jobs

        for cfg in SHEET_CONFIGS:
            sheet_name = cfg["sheet_name"]
            jobs = jobs_by_sheet.get(sheet_name, [])
            if not jobs:
                print(f"{sheet_name}: API:sta ei löytynyt ilmoituksia.", flush=True)
                continue

            print(f"{sheet_name}: API:sta yhteensä {len(jobs)} ilmoitusta.", flush=True)
            df, added, removed_n, _ = sync_dataframe(jobs, sheet_name)
            dfs[sheet_name] = df
            print(
                f"{sheet_name}: synkronointi: +{added} uutta, -{removed_n} poistunutta.",
                flush=True,
            )
            save_excel(df, sheet_name)

            if needs_browser_for_yritys(df):
                print(
                    f"{sheet_name}: täydennetään puuttuvaa Yritys selaimella...",
                    flush=True,
                )
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=True)
                    try:
                        context = browser.new_context(
                            user_agent=(
                                "Mozilla/5.0 (Windows NT 10.0; rv:109.0) "
                                "Gecko/20100101 Firefox/115.0"
                            ),
                        )
                        try:
                            page = context.new_page()
                            fill_missing_yritys_with_browser(page, df, sheet_name)
                        finally:
                            context.close()
                    finally:
                        browser.close()
                save_excel(df, sheet_name)

        koulutus_rows = jobs_by_sheet.get(KOULUTUS_SHEET_NAME, [])
        if koulutus_rows:
            koulutus_df, added_k, removed_k = sync_koulutus_dataframe(
                koulutus_rows, KOULUTUS_SHEET_NAME
            )
            print(
                f"{KOULUTUS_SHEET_NAME}: API:sta yhteensä {len(koulutus_rows)} ohjelmaa.",
                flush=True,
            )
            print(
                f"{KOULUTUS_SHEET_NAME}: synkronointi: +{added_k} uutta, -{removed_k} poistunutta.",
                flush=True,
            )
            save_koulutus_excel(koulutus_df, KOULUTUS_SHEET_NAME)
        else:
            print(f"{KOULUTUS_SHEET_NAME}: API:sta ei löytynyt ohjelmia.", flush=True)

        print(
            "Synkronointi valmis. Tehtävänimike-sarake = hyperlink.",
            flush=True,
        )

    except KeyboardInterrupt:
        print("Keskeytetty.", flush=True)
        for sheet_name, df in dfs.items():
            save_excel(df, sheet_name)
        sys.exit(1)
    except Exception as e:
        print(f"Virhe: {e}", flush=True)
        for sheet_name, df in dfs.items():
            try:
                save_excel(df, sheet_name)
            except Exception:
                pass
        raise


if __name__ == "__main__":
    main()
