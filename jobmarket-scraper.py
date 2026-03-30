#!/usr/bin/env python3
"""
Job collection from tyomarkkinatori.fi into Excel. / Työmarkkinatori: työpaikkojen keruu tyomarkkinatori.fi-sivustolta Exceliin.
List -> sync -> card details -> save with hyperlinks. / Lista -> synkronointi -> ilmoituskorttien tiedot -> tallennus hyperlinkeillä.
"""
import re
import sys
from pathlib import Path
from typing import Optional

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(line_buffering=True)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


# ---------------------------------------------------------------------------
# Configuration / Asetukset
# ---------------------------------------------------------------------------

EXCEL_PATH = Path(__file__).resolve().parent / "tyomarkkinatori_jobs.xlsx"
BASE_DOMAIN = "https://tyomarkkinatori.fi"
BASE_URL = (
    "https://tyomarkkinatori.fi/henkiloasiakkaat/avoimet-tyopaikat"
    "?in=25&or=CLOSING&p={p}&ps=30"
)

PAGE_SIZE = 30
TIMEOUT_MS = 45000
CARD_WAIT_MS = 1500
SAVE_EVERY_N_CARDS = 10

CARD_COLUMNS = [
    "Yritys", "Sijainti", "Työaika", "Palkan peruste", "Työn jatkuvuus",
    "Työskentelyaika", "Julkaistu", "Hakuaika päättyy",
    "Ammatit", "Ammattiryhmä", "Vaadittu koulutustaso", "Ajokortit",
]

# Parsing boundaries for Osaamiset section / Osaamiset-osion jäsennyksen rajat
OSAAMISET_BOUNDARIES = (
    "Ammatit", "Osaamiset", "Vaadittu kielitaito",
    "Vaadittu koulutustaso", "Työssä vaadittavat ajokortit",
)
OSAAMISET_SUBSECTIONS = [
    "Ammatit", "Vaadittu koulutustaso", "Työssä vaadittavat ajokortit",
]
OSAAMISET_TO_COLUMN = {
    "Ammatit": "Ammatit",
    "Vaadittu koulutustaso": "Vaadittu koulutustaso",
    "Työssä vaadittavat ajokortit": "Ajokortit",
}

RE_LISTING = re.compile(r"^https?://[^/]+/henkiloasiakkaat/avoimet-tyopaikat/\?", re.I)
RE_JOB_PATH = re.compile(r"/avoimet-tyopaikat/[^/?]+", re.I)


# ---------------------------------------------------------------------------
# Listing collection / Työpaikkalistan keruu
# ---------------------------------------------------------------------------

def _normalize_url(href: str) -> str:
    """URL without query part and trailing slash. / Verkko-osoite ilman kyselyosaa ja loppuviivaa."""
    return (href or "").split("?")[0].rstrip("/")


def _extract_yritys_from_listing_card(loc, title_hint: str = "") -> str:
    """
    Company on listing card: 'Yritys | Julkaistu …' (walk up from job link).
    / Listauskortilla yritys näkyy rivillä ennen putkea ja Julkaistu-tekstiä.
    """
    hint = (title_hint or "").strip()
    try:
        raw = loc.evaluate(
            """(el, titleHint) => {
                const hint = (titleHint || '').toString().trim();
                const tryLine = (line) => {
                    const s = (line || '').trim();
                    const bar = s.indexOf('|');
                    if (bar < 0) return '';
                    if (!/Julkaistu/i.test(s.slice(bar + 1))) return '';
                    let left = s.slice(0, bar).trim();
                    if (!left || left.length > 200) return '';
                    if (hint && left.startsWith(hint)) {
                        left = left.slice(hint.length).trim();
                        left = left.replace(/^[,|\\s\\u00a0–-]+/, '').trim();
                    }
                    return left;
                };
                let n = el;
                for (let d = 0; d < 14 && n; d++) {
                    const rawText = n.innerText || '';
                    const lines = rawText.split(/\\r?\\n/).map(t => t.trim()).filter(Boolean);
                    for (const line of lines) {
                        const got = tryLine(line);
                        if (got) return got;
                    }
                    const oneLine = rawText.replace(/\\s+/g, ' ').trim();
                    const got = tryLine(oneLine);
                    if (got) return got;
                    n = n.parentElement;
                }
                return '';
            }""",
            hint,
        )
    except Exception:
        return ""
    s = (raw or "").strip()
    return s[:500] if s else ""


def _is_job_link(href: str) -> bool:
    """Job-card link, not listing page link. / Työpaikkailmoituksen korttilinkki, ei listaussivun linkki."""
    if not href or RE_LISTING.match(href):
        return False
    path = (href or "").split("?")[0].rstrip("/")
    if re.search(r"/avoimet-tyopaikat/?$", path):
        return False
    if RE_JOB_PATH.search(href):
        return True
    m = re.search(r"/avoimet-tyopaikat/([^/?]+)", path, re.I)
    if m:
        seg = m.group(1)
        if len(seg) >= 8 and seg.lower() not in ("fi", "sv", "en"):
            return True
    return False


def fetch_listing_page(page, page_num: int, aggressive: bool = False) -> list[dict]:
    """Cards from one listing page. / Yhden listaussivun kortit: Linkki, Tehtävänimike, Yritys (listausrivi)."""
    url = BASE_URL.format(p=page_num)
    for attempt in range(2):
        try:
            page.goto(url, wait_until="networkidle", timeout=TIMEOUT_MS)
            break
        except Exception as e:
            if attempt == 0:
                print(f"  Yritetään uudelleen (sivu {page_num + 1})...", flush=True)
            else:
                print(f"Virhe sivun latauksessa (sivu {page_num + 1}): {e}", flush=True)
                return []

    wait_after_load = 4000 if aggressive else 2500
    scroll_steps = 25 if aggressive else 15
    step_wait = 600 if aggressive else 400
    wait_bottom = 4000 if aggressive else 2500
    try:
        page.wait_for_selector("a[href*='/avoimet-tyopaikat/']", timeout=20000)
        page.wait_for_timeout(wait_after_load)
        for _ in range(2):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1500)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(500)
        for step in range(scroll_steps):
            page.evaluate(f"window.scrollTo(0, {800 * (step + 1)})")
            page.wait_for_timeout(step_wait)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(wait_bottom)
    except Exception:
        pass

    results = []
    seen = set()
    for loc in page.locator("a[href*='/avoimet-tyopaikat/']").all():
        try:
            href = loc.get_attribute("href")
            if not _is_job_link(href):
                continue
            url_norm = _normalize_url(href)
            if url_norm in seen:
                continue
            seen.add(url_norm)
            title = (loc.inner_text() or "").strip()
            if not title or len(title) > 500:
                title = loc.get_attribute("title") or ""
            title = title or "-"
            yritys = _extract_yritys_from_listing_card(loc, title_hint=title)
            yritys = _clean_company_name(yritys)
            if _looks_like_location(yritys):
                yritys = ""
            results.append({"Linkki": url_norm, "Tehtävänimike": title, "Yritys": yritys})
        except Exception:
            continue

    if not results:
        n = len(page.locator("a[href*='/avoimet-tyopaikat/']").all())
        if n:
            print(f"  (Huom: {n} linkkiä, mikään ei täsmännyt.)", flush=True)
    return results


def fetch_all_listings(page) -> list[dict]:
    """All jobs from paginated listings. / Kaikki työpaikat sivutetusta listasta."""
    all_jobs = []
    page_num = 0
    same_page_fail = 0
    max_empty_retries = 4
    while True:
        print(f"Haetaan sivua {page_num + 1}...", flush=True)
        use_aggressive = page_num == 7
        batch = fetch_listing_page(page, page_num, aggressive=use_aggressive)
        expected = PAGE_SIZE
        if batch and len(batch) < expected:
            print(f"  Sivu {page_num + 1}: vain {len(batch)} korttia (odotus {expected}), haetaan uudelleen...", flush=True)
            batch2 = fetch_listing_page(page, page_num, aggressive=True)
            if len(batch2) > len(batch):
                batch = batch2
                print(f"  Uudelleenhaulla {len(batch)} korttia.", flush=True)
        if not batch:
            same_page_fail += 1
            if same_page_fail >= max_empty_retries:
                print(f"  Lopetetaan (sivu {page_num + 1} tyhjä {same_page_fail} kertaa).", flush=True)
                break
            if len(all_jobs) % PAGE_SIZE == 0 and len(all_jobs) >= PAGE_SIZE:
                max_empty_retries = 6
            print(f"  Tyhjä sivu, yritetään uudelleen ({same_page_fail}/{max_empty_retries})...", flush=True)
            page.wait_for_timeout(4000)
            continue
        same_page_fail = 0
        max_empty_retries = 4
        seen = {j["Linkki"] for j in all_jobs}
        for job in batch:
            if job["Linkki"] not in seen:
                seen.add(job["Linkki"])
                all_jobs.append(job)
        print(f"Löytyi {len(batch)} ilmoitusta (yhteensä {len(all_jobs)}).", flush=True)
        if len(batch) < PAGE_SIZE:
            break
        page_num += 1
    return all_jobs


# ---------------------------------------------------------------------------
# Card text parsing / Ilmoituskortin tekstin jäsennys
# ---------------------------------------------------------------------------

def _extract_value(
    text: str,
    label: str,
    stop_headers: list[str],
    max_len: int = 600,
    section: Optional[str] = None,
) -> str:
    """
Value after label until first stop header. / Arvo otsakkeen jälkeen ensimmäiseen stop_header-arvoon asti.
section: search label only in this text fragment. / section: etsi otsake vain tästä tekstikatkelmasta.
    """
    search = (section or text) or ""
    lines = search.split("\n")
    label_esc = re.escape(label)
    for i, line in enumerate(lines):
        ls = line.strip()
        if not (
            ls == label
            or (ls.startswith(label) and (len(ls) == len(label)
                or ls[len(label):].strip().startswith(":")))
        ):
            continue
        vals = []
        for j in range(i + 1, len(lines)):
            r = lines[j].strip()
            if not r:
                continue
            if any(h in r or r.startswith(h.strip()) for h in stop_headers):
                break
            vals.append(r)
        if vals:
            return " ".join(vals).strip()[:max_len]
        return ""
    return ""


def _extract_next_line(text: str, label: str, max_len: int = 500) -> str:
    """One line after label (regex). / Yksi rivi otsakkeen jälkeen (säännöllinen lauseke)."""
    m = re.search(re.escape(label) + r"\s*\n\s*([^\n]+)", text)
    return m.group(1).strip()[:max_len] if m else ""


def _parse_osaamiset_section(text: str) -> dict:
    """Parse the Tehtävään liittyvät taidot ja osaamiset section. / Jäsennä osio Tehtävään liittyvät taidot ja osaamiset."""
    out = {col: "" for col in OSAAMISET_TO_COLUMN.values()}
    if not text:
        return out
    idx = text.find("Tehtävään liittyvät taidot ja osaamiset")
    if idx < 0:
        return out
    end = text.find("Työpaikan sijainti", idx)
    if end < 0:
        end = text.find("Muut tiedot", idx)
    if end < 0:
        end = len(text)
    block = text[idx:end]
    lines = [l.strip() for l in block.split("\n") if l.strip()]
    for header in OSAAMISET_SUBSECTIONS:
        col = OSAAMISET_TO_COLUMN.get(header)
        if not col:
            continue
        for j, line in enumerate(lines):
            if line != header:
                continue
            vals = []
            for k in range(j + 1, len(lines)):
                nxt = lines[k]
                if nxt in OSAAMISET_BOUNDARIES or nxt in (
                    "Työpaikan sijainti", "Muut tiedot", "Tarkemmat tiedot",
                ):
                    break
                vals.append(nxt)
            if vals:
                out[col] = " ".join(vals).strip()[:600]
            break
    return out


def _parse_card_text(text: str) -> dict:
    """Parse all card fields from main.innerText. / Jäsennä kaikki kortin kentät main.innerText-arvosta."""
    out = {k: "" for k in CARD_COLUMNS}
    if not text:
        return out

    m = re.search(r"Yritys\s*[:\s\xa0]+\s*([^\n]+)", text)
    if m:
        out["Yritys"] = m.group(1).strip()[:500]

    m = re.search(r"Työpaikan\s+sijainti\s*\n\s*([^\n]+)", text, re.I)
    if m:
        first = m.group(1).strip()
        if first.lower() in ("sijainti", "sijainti :", "sijainti:") or (
            first.lower().startswith("sijainti") and len(first) < 20
        ):
            m2 = re.search(
                r"Työpaikan\s+sijainti\s*\n\s*[^\n]+\s*\n\s*[:\s\xa0]*\s*([^\n]+)",
                text, re.I,
            )
            if m2:
                out["Sijainti"] = m2.group(1).strip()[:500]
        else:
            out["Sijainti"] = first[:500]
    if not out["Sijainti"]:
        m = re.search(r"Sijainti\s*[:\s\xa0]+\s*([^\n]+)", text)
        if m:
            sij = m.group(1).strip()
            for stop in (
                "Kokoaikatyö", "Osa-aikatyö", "Pääsääntöisesti", "Työaika",
                "Työn jatkuvuus", "Työskentelyaika", "Tarkemmat", "klo ",
            ):
                if stop in sij:
                    sij = sij.split(stop)[0].strip().rstrip(",")
                    break
            out["Sijainti"] = sij[:500]

    tarkemmat = text[text.find("Tarkemmat tiedot"):] if "Tarkemmat tiedot" in text else text
    out["Työaika"] = (
        _extract_value(text, "Työaika", ["Työn jatkuvuus", "Työ alkaa"], 500, tarkemmat)
        or _extract_value(text, "Työaika", ["Työn jatkuvuus", "Työ alkaa"])
        or _extract_next_line(text, "Työaika")
    )
    out["Palkan peruste"] = (
        _extract_value(text, "Palkan peruste", ["Työskentelyaika", "Työ alkaa", "Avointen", "Lisätietoja"])
        or _extract_next_line(text, "Palkan peruste")
    )
    out["Työn jatkuvuus"] = (
        _extract_value(text, "Työn jatkuvuus", ["Työ alkaa", "Palkan peruste", "Lisätietoja"])
        or _extract_next_line(text, "Työn jatkuvuus")
    )
    out["Työskentelyaika"] = _extract_next_line(text, "Työskentelyaika")

    m = re.search(r"Julkaistu\s+([0-9]{1,2}\.[0-9]{1,2}\.[0-9]{4})", text)
    if m:
        out["Julkaistu"] = m.group(1).strip()
    m = re.search(r"Hakuaika\s+päättyy\s*[:\s\xa0]+\s*([0-9]{1,2}\.[0-9]{1,2}\.[0-9]{4})", text)
    if m:
        out["Hakuaika päättyy"] = m.group(1).strip()

    for col, val in _parse_osaamiset_section(text).items():
        if val and col in out:
            out[col] = val

    m = re.search(r"Ammatti(?:nimike)?\s*[:\s\xa0]+\s*([^\n]+)", text, re.I)
    if m:
        out["Ammatit"] = m.group(1).strip()[:500]
    if not out["Ammatit"]:
        out["Ammatit"] = _extract_next_line(text, "Ammatti")

    m = re.search(
        r"Ammattiryhmä\s*[:\s\xa0]+\s*([^\n]+(?:>\s*[^\n]+)*)",
        text, re.I
    )
    if m:
        path = " > ".join(s.strip() for s in m.group(1).split(">"))
        out["Ammattiryhmä"] = path[:500]

    return out


def _clean_company_name(value: str) -> str:
    """Return a plausible company name or empty string. / Palauta kelvollinen yritysnimi tai tyhjä merkkijono."""
    s = (value or "").strip()
    if not s:
        return ""
    # Filter out long description-like sentences.
    if len(s) > 80:
        return ""
    low = s.lower()
    bad_tokens = (" toimii ", " yrityksessä", " tehtävässä", " rooli ", " tiimi ")
    if any(t in low for t in bad_tokens):
        return ""
    return s


def _looks_like_location(value: str) -> bool:
    """Detect location-like values in Yritys. / Tunnista sijaintia muistuttavat Yritys-arvot."""
    s = (value or "").strip()
    if not s:
        return False
    low = s.lower()
    location_tokens = (
        "helsinki", "espoo", "vantaa", "tampere", "turku", "hämeenlinna",
        "koko suomi", "suomi", "tai", "etätyö", "etatyö",
    )
    if any(tok in low for tok in location_tokens):
        return True
    # Single short token is often a city name in this dataset.
    if " " not in s and len(s) <= 14:
        return True
    return False


def _company_from_title(title: str) -> str:
    """Conservative company fallback from title. / Varovainen yritys-fallback otsikosta."""
    t = (title or "").strip()
    if not t or "," not in t:
        return ""
    cand = t.split(",")[-1].strip()
    cand = _clean_company_name(cand)
    if not cand or _looks_like_location(cand):
        return ""
    low = cand.lower()
    company_tokens = (" oy", " oyj", " ab", " ltd", " ky", " tmi", " ry", " inc")
    if any(tok in low for tok in company_tokens):
        return cand
    return ""


def _extract_company_from_page(page) -> str:
    """Extract company directly from vacancy page. / Poimi yritys suoraan työpaikkasivulta."""
    try:
        value = page.evaluate(
            """
            () => {
                const clean = (v) => (v || '').toString().trim();

                // 1) JSON-LD (JobPosting.hiringOrganization.name)
                const ldScripts = Array.from(document.querySelectorAll('script[type="application/ld+json"]'));
                for (const s of ldScripts) {
                    try {
                        const raw = clean(s.textContent);
                        if (!raw) continue;
                        const data = JSON.parse(raw);
                        const items = Array.isArray(data) ? data : [data];
                        for (const item of items) {
                            const org = item && item.hiringOrganization;
                            const name = clean(org && org.name);
                            if (name) return name;
                        }
                    } catch (_) {}
                }

                // 2) Label-based extraction from visible text
                const bodyText = clean(document.body && document.body.innerText);
                if (bodyText) {
                    const m = bodyText.match(/Yritys\\s*[:\\s\\u00a0]+([^\\n]+)/i);
                    if (m && m[1]) return clean(m[1]);
                }

                // 3) Common company selectors
                const selectors = [
                    '[data-testid*="company"]',
                    '[class*="company"]',
                    '[id*="company"]'
                ];
                for (const sel of selectors) {
                    const el = document.querySelector(sel);
                    const t = clean(el && el.textContent);
                    if (t) return t;
                }
                return '';
            }
            """
        )
        return (value or "").strip()
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Card loading / Ilmoituskortin lataus
# ---------------------------------------------------------------------------

def _extract_ammattiryhma(page) -> str:
    """Extract Ammattiryhmä path from page. / Poimi Ammattiryhmä-polku (Erityisasiantuntijat > ... > Sovellusohjelmoijat) sivulta."""
    try:
        path = page.evaluate("""
            () => {
                const text = document.body ? document.body.innerText : '';
                const markers = ['Ammattiryhmä', 'Ammattiala', 'Ammattiluokitus'];
                for (const m of markers) {
                    const idx = text.indexOf(m);
                    if (idx >= 0) {
                        const rest = text.substring(idx + m.length).replace(/^[:\\s\\xa0]+/, '');
                        const line = rest.split('\\n')[0].trim();
                        if (line.indexOf(' > ') >= 0) return line.replace(/\\s+/g, ' ').substring(0, 500);
                        const match = rest.match(/^([^\\n]+?)(?=\\n[A-ZÄÖÅ]|$)/);
                        if (match) return match[1].trim().substring(0, 500);
                    }
                }
                const lines = text.split('\\n');
                for (const line of lines) {
                    const t = line.trim();
                    if (t.split(' > ').length >= 3 && /^[A-Za-zäöåÄÖÅ0-9\\-\\s]+$/.test(t)) {
                        return t.substring(0, 500);
                    }
                }
                return '';
            }
        """)
        return (path or "").strip()
    except Exception:
        return ""


def fetch_card_details(page, url: str, title_hint: str = "") -> dict:
    """Open card and parse all fields. / Avaa kortti ja jäsennä kaikki kentät."""
    out = {k: "" for k in CARD_COLUMNS}
    s = (url or "").strip()
    full_url = s if s.startswith("http") else f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    try:
        page.goto(full_url, wait_until="load", timeout=TIMEOUT_MS)
        page.wait_for_timeout(CARD_WAIT_MS)
    except Exception:
        return out

    try:
        btn = page.locator("text=Tehtävään liittyvät taidot ja osaamiset").first
        if btn.count() > 0:
            btn.click()
            page.wait_for_timeout(800)
    except Exception:
        pass

    try:
        main_el = page.locator("main").first
        if main_el.count() > 0:
            out = _parse_card_text(main_el.inner_text())
    except Exception:
        pass

    if not out.get("Ammattiryhmä"):
        out["Ammattiryhmä"] = _extract_ammattiryhma(page)
    if not out.get("Yritys"):
        out["Yritys"] = _extract_company_from_page(page)
    # Keep only plausible company names.
    out["Yritys"] = _clean_company_name(out.get("Yritys", ""))
    if not out["Yritys"]:
        out["Yritys"] = _company_from_title(title_hint)
    return out


# ---------------------------------------------------------------------------
# Excel: sync and save / Excel: synkronointi ja tallennus
# ---------------------------------------------------------------------------

def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Linkki", "Tehtävänimike"] + CARD_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def _extract_urls_and_titles_from_excel(path: Path, n_rows: int) -> tuple[list[str], list[str]]:
    """URL and display text from Tehtävänimike hyperlinks/formulas. / Verkko-osoite ja näkyvä teksti Tehtävänimike-sarakkeen linkeistä/HYPERLINK-kaavoista."""
    urls: list[str] = []
    titles: list[str] = []
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb.active
        col_idx = 1
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Tehtävänimike":
                col_idx = c
                break
        for r in range(2, ws.max_row + 1):
            if len(urls) >= n_rows:
                break
            cell = ws.cell(row=r, column=col_idx)
            href = None
            title = ""
            val = cell.value
            h = getattr(cell, "hyperlink", None)
            if h:
                href = getattr(h, "target", None) or getattr(h, "location", None)
            if isinstance(val, str) and "HYPERLINK" in val:
                m_url = re.search(r'HYPERLINK\s*\(\s*"((?:[^"]|"")+)"', val, re.I)
                if m_url:
                    href = href or m_url.group(1).replace('""', '"')
                m_title = re.search(r'HYPERLINK\s*\(\s*"(?:[^"]|"")*"\s*,\s*"((?:[^"]|"")*)"\s*\)', val, re.I)
                if m_title:
                    title = m_title.group(1).replace('""', '"').strip()
            elif isinstance(val, str) and val.strip() and not val.strip().startswith("="):
                title = val.strip()[:500]
            urls.append(str(href or "").strip())
            titles.append(title)
        wb.close()
    except Exception:
        pass
    return urls, titles


def _norm_link(lk: str) -> str:
    s = str(lk).strip()
    if not s or s == "nan":
        return ""
    s = s.split("?")[0].rstrip("/")
    if s and not s.startswith("http"):
        s = f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    return s


def sync_dataframe(jobs_from_web: list[dict]) -> tuple[pd.DataFrame, int, int, set[str]]:
    """
Table contains ONLY jobs currently on the site. Add new, remove missing. / Taulukko = VAIN sivustolla olevat työpaikat. Lisää uudet, poista puuttuvat.
Guarantee: len(df) == len(jobs_from_web). / Takuu: len(df) == len(jobs_from_web).
    """
    cols = ["Linkki", "Tehtävänimike"] + CARD_COLUMNS
    web_links = {_norm_link(j["Linkki"]) for j in jobs_from_web}
    web_links.discard("")

    if not EXCEL_PATH.exists():
        df = pd.DataFrame(jobs_from_web)
        df = _ensure_columns(df)
        return df, len(jobs_from_web), 0, set()

    try:
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        print(f"Excelin lukuvirhe: {e}", flush=True)
        df = pd.DataFrame(jobs_from_web)
        df = _ensure_columns(df)
        return df, len(jobs_from_web), 0, set()

    df = _ensure_columns(df)
    if "Tehtävänimike" in df.columns and df["Tehtävänimike"].dtype != object:
        df["Tehtävänimike"] = df["Tehtävänimike"].astype(object)
    urls, titles = _extract_urls_and_titles_from_excel(EXCEL_PATH, len(df))
    if len(urls) < len(df):
        urls = urls + [""] * (len(df) - len(urls))
        titles = titles + [""] * (len(df) - len(titles))
    df["Linkki"] = urls[: len(df)]
    def _is_empty_title(x) -> bool:
        if pd.isna(x):
            return True
        s = str(x).strip().lower()
        return not s or s == "nan"

    for i, tit in enumerate(titles[: len(df)]):
        if tit and tit.lower() != "nan" and _is_empty_title(df.at[i, "Tehtävänimike"]):
            df.at[i, "Tehtävänimike"] = tit

    excel_links = {_norm_link(x) for x in df["Linkki"].astype(str).unique() if _norm_link(x)}
    new_links = web_links - excel_links
    removed_links = excel_links - web_links

    df = df[~df["Linkki"].astype(str).apply(_norm_link).isin(removed_links)].copy()
    new_rows = [j for j in jobs_from_web if _norm_link(j["Linkki"]) in new_links]
    if new_rows:
        new_df = pd.DataFrame(new_rows)
        for c in df.columns:
            if c not in new_df.columns:
                new_df[c] = ""
        new_df = new_df.reindex(columns=df.columns, fill_value="")
        df = pd.concat([df, new_df], ignore_index=True)

    df = df[df["Linkki"].astype(str).apply(_norm_link).isin(web_links)].copy()
    df = df.drop_duplicates(subset=["Linkki"], keep="first")
    link_to_title = {_norm_link(j["Linkki"]): j.get("Tehtävänimike", "") for j in jobs_from_web}
    for i in df.index:
        if _is_empty_title(df.at[i, "Tehtävänimike"]):
            t = link_to_title.get(_norm_link(str(df.at[i, "Linkki"])), "-")
            if t and str(t).strip().lower() != "nan":
                df.at[i, "Tehtävänimike"] = str(t)

    link_to_yritys: dict[str, str] = {}
    for j in jobs_from_web:
        lk = _norm_link(j["Linkki"])
        if not lk:
            continue
        raw_y = str(j.get("Yritys", "") or "").strip()
        if not raw_y:
            continue
        y = _clean_company_name(raw_y)
        if y and not _looks_like_location(y):
            link_to_yritys[lk] = y
    for i in df.index:
        y = link_to_yritys.get(_norm_link(str(df.at[i, "Linkki"])), "")
        if y:
            df.at[i, "Yritys"] = y

    order = {_norm_link(j["Linkki"]): i for i, j in enumerate(jobs_from_web)}
    df["_ord"] = df["Linkki"].astype(str).apply(lambda x: order.get(_norm_link(x), 999))
    df = df.sort_values("_ord").drop(columns=["_ord"])
    df.reset_index(drop=True, inplace=True)
    return df, len(new_rows), len(removed_links), removed_links


def _is_row_complete(df: pd.DataFrame, i: int) -> bool:
    """Row is complete when Yritys is filled. / Rivi on valmis, kun Yritys on täytetty."""
    if not (
        "Yritys" in df.columns
        and df.at[i, "Yritys"] is not None
        and str(df.at[i, "Yritys"]).strip() != ""
    ):
        return False
    return not _looks_like_location(str(df.at[i, "Yritys"]))


def _get_url_from_hyperlink(cell) -> Optional[str]:
    """Extract URL from cell hyperlink. / Poimi verkko-osoite solun hyperlinkistä."""
    h = getattr(cell, "hyperlink", None)
    if h:
        return getattr(h, "target", None) or getattr(h, "location", None)
    if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK("):
        m = re.search(r'=HYPERLINK\s*\(\s*"((?:[^"]|"")+)"', cell.value)
        if m:
            return m.group(1).replace('""', '"')
    return None


def _cols_and_title_idx(ws) -> tuple[list[str], int]:
    """Our columns list and Tehtävänimike column index. / Omien sarakkeiden lista ja Tehtävänimike-sarakkeen indeksi."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    title_col = 1
    for c, val in enumerate(header, 1):
        if val == "Tehtävänimike":
            title_col = c
            break
    our_cols = ["Tehtävänimike", "Yritys"]
    return our_cols, title_col


def save_excel(
    df: pd.DataFrame,
    removed_links: Optional[set[str]] = None,
) -> None:
    """
    Save to Excel without rewriting the whole sheet; preserve order and formatting,
    add only new rows, and remove outdated job rows.
    / Tallennus Exceliin ilman koko taulukon ylikirjoitusta; säilytä järjestys ja
    muotoilu, lisää vain uudet rivit ja poista vanhentuneet työpaikkarivit.
    """
    df = _ensure_columns(df)
    removed_links = removed_links or set()
    our_cols = ["Tehtävänimike", "Yritys"]

    if not EXCEL_PATH.exists():
        df_out = df.reindex(columns=our_cols, fill_value="")
        df_out = df_out[our_cols]
        df_out.to_excel(EXCEL_PATH, index=False, engine="openpyxl")
        _apply_hyperlinks(EXCEL_PATH, df["Linkki"].astype(str).tolist())
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    existing_cols, title_col = _cols_and_title_idx(ws)

    def _norm(lk: str) -> str:
        s = str(lk).strip()
        if not s or s == "nan":
            return ""
        s = s.split("?")[0].rstrip("/")
        if s and not s.startswith("http"):
            s = f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
        return s

    linkki_to_row = {}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=title_col)
        url = _get_url_from_hyperlink(cell)
        if url:
            nurl = _norm(url)
            if nurl:
                linkki_to_row[nurl] = r

    df_linkit = {_norm(str(row.get("Linkki", ""))) for _, row in df.iterrows()}
    df_linkit.discard("")
    keep_rows = {r for url, r in linkki_to_row.items() if url in df_linkit}
    to_delete = sorted((r for r in range(2, ws.max_row + 1) if r not in keep_rows), reverse=True)
    if to_delete:
        print(f"Poistetaan {len(to_delete)} riviä (vanhentuneet + tyhjät).", flush=True)
    for row in to_delete:
        ws.delete_rows(row, 1)
        for k in list(linkki_to_row):
            if linkki_to_row[k] > row:
                linkki_to_row[k] -= 1
            elif linkki_to_row[k] == row:
                del linkki_to_row[k]

    col_indices = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h in our_cols:
            col_indices[h] = c
    for col in our_cols:
        if col not in col_indices:
            new_c = ws.max_column + 1
            ws.cell(row=1, column=new_c, value=col)
            col_indices[col] = new_c

    for _, row in df.iterrows():
        linkki_raw = str(row.get("Linkki", "")).strip()
        if not linkki_raw or linkki_raw == "nan":
            continue
        linkki = _norm(linkki_raw)
        if not linkki:
            continue

        if linkki in linkki_to_row:
            excel_row = linkki_to_row[linkki]
            for col in our_cols:
                if col == "Tehtävänimike":
                    continue
                if col in col_indices and col in row:
                    val = row[col]
                    if pd.notna(val) and str(val).strip():
                        ws.cell(row=excel_row, column=col_indices[col]).value = str(val)[:500]
        else:
            new_row = ws.max_row + 1
            for col in our_cols:
                if col in col_indices and col in row:
                    val = row[col] if col != "Tehtävänimike" else row.get("Tehtävänimike", "-")
                    if pd.notna(val) and str(val).strip():
                        ws.cell(row=new_row, column=col_indices[col]).value = str(val)[:500]
            _set_hyperlink(ws, new_row, title_col, linkki, display=str(row.get("Tehtävänimike", "-")))
            linkki_to_row[linkki] = new_row

    _apply_hyperlinks_to_ws(ws, df, title_col, our_cols, col_indices)
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=title_col).value:
            last_data_row = r
    while ws.max_row > last_data_row:
        ws.delete_rows(ws.max_row, 1)
    if ws.max_column >= 1:
        ws.column_dimensions["A"].width = 60
    if ws.max_column >= 2:
        ws.column_dimensions["B"].width = 80
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    wb.save(EXCEL_PATH)
    n_rows = ws.max_row - 1
    wb.close()
    print(f"Tallennettu: {n_rows} riviä.", flush=True)


def _apply_hyperlinks(path: Path, urls: list[str]) -> None:
    """Add hyperlinks to Tehtävänimike in a new file. / Lisää hyperlinkit Tehtävänimike-sarakkeeseen uudessa tiedostossa."""
    try:
        wb = load_workbook(path)
        ws = wb.active
        title_col = 1
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Tehtävänimike":
                title_col = c
                break
        for r in range(2, min(ws.max_row + 1, len(urls) + 2)):
            idx = r - 2
            if idx < len(urls) and urls[idx]:
                _set_hyperlink(ws, r, title_col, urls[idx])
        wb.save(path)
        wb.close()
    except Exception:
        pass


def _set_hyperlink(ws, row: int, col: int, url: str, display: Optional[str] = None) -> None:
    """Set hyperlink to cell. / Aseta hyperlinkki soluun."""
    s = str(url).strip()
    if ".fihenkilo" in s:
        s = s.replace(".fihenkilo", ".fi/henkilo", 1)
    elif not s.startswith("http"):
        s = f"{BASE_DOMAIN.rstrip('/')}/{s.lstrip('/')}"
    cell = ws.cell(row=row, column=col)
    raw = (display or cell.value or s)
    label = ("-" if (raw is None or str(raw).strip().lower() in ("", "nan")) else str(raw)).replace('"', '""')
    url_esc = s.replace('"', '""')
    cell.value = f'=HYPERLINK("{url_esc}", "{label}")'
    cell.font = Font(color="0563C1", underline="single")


def _apply_hyperlinks_to_ws(ws, df: pd.DataFrame, title_col: int, our_cols: list, col_indices: dict) -> None:
    """Update hyperlinks for all rows by Linkki (URL + display text). / Päivitä hyperlinkit kaikille riveille Linkki-arvon mukaan (verkko-osoite + näkyvä teksti)."""
    linkki_to_row = {}
    for _, r in df.iterrows():
        lk = _norm_link(str(r.get("Linkki", "")))
        if lk:
            linkki_to_row[lk] = r
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=title_col)
        url = _get_url_from_hyperlink(cell)
        if url:
            nurl = _norm_link(url)
            if nurl in linkki_to_row:
                row_data = linkki_to_row[nurl]
                disp = row_data.get("Tehtävänimike")
                disp = "-" if (pd.isna(disp) or not str(disp).strip()) else str(disp)
                _set_hyperlink(ws, r, title_col, url, display=disp)


def fill_card_details(page, df: pd.DataFrame) -> None:
    """Fill Yritys for rows where it is missing. / Täytä Yritys riveille, joilta se puuttuu."""
    if "Yritys" in df.columns:
        for i in range(len(df)):
            if _looks_like_location(str(df.at[i, "Yritys"])):
                df.at[i, "Yritys"] = ""

    n = len(df)
    saved = 0
    skipped = 0
    for i in range(n):
        link = df.at[i, "Linkki"]
        if pd.isna(link) or not str(link).strip():
            continue
        if _is_row_complete(df, i):
            skipped += 1
            continue
        print(f"Haetaan kortti {i + 1}/{n}...", flush=True)
        try:
            title_hint = str(df.at[i, "Tehtävänimike"]) if "Tehtävänimike" in df.columns else ""
            data = fetch_card_details(page, str(link).strip(), title_hint=title_hint)
            val = data.get("Yritys", "")
            if val:
                df.at[i, "Yritys"] = val
                saved += 1
        except Exception as e:
            print(f"Kortti {i + 1}/{n} ohitettu: {e}", flush=True)

        if saved > 0 and saved % SAVE_EVERY_N_CARDS == 0:
            print(f"Varmuuskopio ({saved} korttia).", flush=True)
            save_excel(df)

    if skipped:
        print(f"Ohitettiin {skipped} valmiiksi täytettyä korttia.", flush=True)


# ---------------------------------------------------------------------------
# Main loop / Pääsilmukka
# ---------------------------------------------------------------------------

def main() -> None:
    print("Työmarkkinatori -synkronointi alkaa.", flush=True)
    df = None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; rv:109.0) Gecko/20100101 Firefox/115.0",
            )
            page = context.new_page()

            try:
                jobs = fetch_all_listings(page)
                if not jobs:
                    print("Verkosta ei löytynyt ilmoituksia.", flush=True)
                    return
                print(f"Verkosta yhteensä {len(jobs)} ilmoitusta.", flush=True)

                df, added, removed, removed_links = sync_dataframe(jobs)
                print(f"Synkronointi: +{added} uutta, -{removed} poistunutta.", flush=True)
                save_excel(df, removed_links=removed_links)

                print("Haetaan Yritys-tiedot...", flush=True)
                fill_card_details(page, df)

                save_excel(df, removed_links=removed_links)
                print("Synkronointi valmis. Tehtävänimike-sarake = hyperlink.", flush=True)
            finally:
                try:
                    context.close()
                except Exception:
                    pass
                try:
                    browser.close()
                except Exception:
                    pass
    except KeyboardInterrupt:
        print("Keskeytetty.", flush=True)
        if df is not None:
            save_excel(df)
        sys.exit(1)
    except Exception as e:
        print(f"Virhe: {e}", flush=True)
        if df is not None:
            try:
                save_excel(df)
            except Exception:
                pass
        raise


if __name__ == "__main__":
    main()
